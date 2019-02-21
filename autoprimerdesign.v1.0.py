#!/usr/bin/python

# ------------------------------------
# python modules
# ------------------------------------
# GSA environment
import os,sys,subprocess
import sys
from optparse import OptionParser
import logging
import time
import json
#import primer3
#from ensemblrest import EnsemblRest
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import pybedtools
#from Bio import Entrez, SeqIO
# ------------------------------------
# constants
# ------------------------------------
# Report errors and warnings while running the script
error   = logging.critical      
warn    = logging.warning

# ------------------------------------
# Misc functions
# ------------------------------------

def primerParser(para,name,seqid):
    '''The main function to call the primer design program.
    primer3_core parameter.txt --output output.txt
    '''
    if name==seqid:
        outputName=name+'_primerDesign_results_output.txt'
    if name!=seqid:
        outputName=name+'_'+seqid+'_primerDesign_results_output.txt'
    cmd='primer3_core '+para+' --output '+outputName
    print("Primer Design Main: "+cmd)
    os.system(cmd)

def primerDesign(input_seq,para,name):
    '''
    add parameters of sequence_template=ATCG and sequence_target=start,length
    '''
    start=input_seq.index(']')+1

    length=input_seq.index(']')-input_seq.index('[')-1
    print('SNP location:'+str(start)+';'+str(length))
    input_seq=input_seq.replace('[','').replace(']','')
    
    #SEQUENCE_TEMPLATE
    string1='SEQUENCE_TEMPLATE='+input_seq
    #SEQUENCE_TARGET
    string2='SEQUENCE_TARGET='+str(start)+','+str(length)

    
    paranew=para.split('.txt')[0]+'_'+name+'.txt'
#    cmd0='cp '+para+' '+paranew
#    print('Create a new parameter input file: '+cmd0)
#    os.system(cmd0)
    para_old=open(para,'r')
    para_new=open(paranew,'w')
    for line in para_old:
        print(line)
        if line=='\n':
            continue
        print(line)
        para_new.writelines(line)
    #parameters=open(paranew,'a')
    para_new.writelines(string1+'\n'+string2+'\n=')
    para_new.close()
    
    return(para_new)

def fetchSequence(chrName,start,end,fastq):
    position='  '.join([chrName,start,end])
    #print(position)
    bedpos = pybedtools.BedTool(position,from_string=True)
    fasta=pybedtools.example_filename(fastq)
    bedpos=bedpos.sequence(fi=fasta)
    sequence=open(bedpos.seqfn).read()
    return(sequence)


def append_df_to_excel(filename, df, sheet_name, startrow=None,truncate_sheet=False, **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    #from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def primerDesign2(start,snplen,seqs,para,name,adeptor,data):
    '''
    step1: add parameters of sequence_template=ATCG and sequence_target=start,length into new parameter files
    step2: call primer3_core function to design the primer
    '''

    #SEQUENCE_TEMPLATE
    string1='SEQUENCE_TEMPLATE='+seqs
    #SEQUENCE_TARGET
    string2='SEQUENCE_TARGET='+str(start)+','+str(snplen)
    #path=os.getcwd()
    paranew=name+'_'+para.split('/')[-1].split('.txt')[0]+'.txt'
    cmd0='cp '+para+' '+paranew
    print('Create a new parameter input file: '+cmd0)
    os.system(cmd0)
    parameters=open(paranew,'a')
    parameters.writelines(string1+'\n'+string2+'\n=')
    parameters.close()
    
    # step2: call primer3_core function
    cmd2='primer3_core '+paranew
    
    primerResults=subprocess.Popen(cmd2,shell=True,stdout=subprocess.PIPE).stdout.readlines()

    result={}
    for n in range(0,5):
        result['OLIGO_'+str(n)]={'LEFT':{},'RIGHT':{}}

    #print(result)
    indexs='No'
    for j in primerResults:
        j=j.decode()
        if j.startswith('SEQUENCE_TARGET'):
            indexs='Yes'
            continue

        if not indexs=='Yes':
            continue
        paras=j[:-1].split('=')[0]
        values=j[:-1].split('=')[1]
        
        if len(paras.split('_'))<3:
            continue
        if paras.split('_')[2].isdigit() and paras.split('_')[1]=='LEFT':

            num=paras.split('_')[2]
            paras=paras.split('_')
                    #print(paras)
            paras.remove(num)
                    #print(paras)
            paras.remove('LEFT')
            subpara='_'.join(paras)
                    #print(subpara)
            if subpara.endswith('SEQUENCE'):
                result['OLIGO_'+num]['LEFT'][subpara]=values
                result['OLIGO_'+num]['LEFT']['primer']=adeptor[0]+values
                continue
            result['OLIGO_'+num]['LEFT'][subpara]=values
            result['OLIGO_'+num]['LEFT']['adeptor_name']='F'+num+'M13'
            continue

        if paras.split('_')[2].isdigit() and paras.split('_')[1]=='RIGHT':
            num=paras.split('_')[2]
            paras=paras.split('_')
                    #print(paras)
            paras.remove(num)
                    #print(paras)
            paras.remove('RIGHT')
            subpara='_'.join(paras)
                    #print(subpara)
            if subpara.endswith('SEQUENCE'):
                result['OLIGO_'+num]['RIGHT'][subpara]=values
                result['OLIGO_'+num]['RIGHT']['primer']=adeptor[1]+values
                continue
            result['OLIGO_'+num]['RIGHT'][subpara]=values
            result['OLIGO_'+num]['RIGHT']['adeptor_name']='R'+num+'M13'
            continue
        #print(paras)
        if paras.startswith('PRIMER_PAIR') and paras.endswith('PRODUCT_SIZE'):
            product_size=values
            num=paras.split('_')[2]
            result['OLIGO_'+num]['LEFT']['size']=product_size
            result['OLIGO_'+num]['RIGHT']['size']=product_size
    print(num)
    #print(pd.DataFrame(result['OLIGO_0']))        
    #print(result)
    # convert to pandas dataframe and output
    #output_wb=Workbook()
    #filepath=path+'/'+name+'_primer_design_output.xlsx'
    #output_wb.save(filepath)

    #filepath_txt=path+'/'+name+'_primer_design_output.txt'
    #output_txt=open(filepath_txt,'w')

    #sheetName=time.strftime("%Y%m%d",time.localtime(time.time()))
    num=int(num)+1
    primer_result_pd=pd.DataFrame()
    for k in range(0,num):
        df=pd.DataFrame(result['OLIGO_'+str(k)])
        df=df.T
        print(df)
        pos=pd.DataFrame(df['PRIMER'].str.split(',',1).tolist(),columns=['START','LEN'])
        pos.index=['LEFT','RIGHT']

                #print(pos)
        datanew=data[['RUNNAME','SAMPLE_ID','CHROM','POS']].applymap(str)
        #print(datanew)
        datanew['LOCI']=datanew[['CHROM','POS']].apply(lambda x: 'chr'+':'.join(x),axis=1)
        datanew=pd.concat([datanew,datanew],axis=0)
        #print(data)
        datanew.index=['LEFT','RIGHT']
        #print(datanew)
        #print(df.to_string())


        dfnew=pd.concat([df[['primer','size']],pos[['START','LEN']],df[['PRIMER_TM','PRIMER_GC_PERCENT','PRIMER_SELF_ANY_TH','PRIMER_SELF_END_TH','PRIMER_HAIRPIN_TH','PRIMER_SEQUENCE','adeptor_name']],datanew[['RUNNAME','SAMPLE_ID','LOCI']]],axis=1)
        dfnew.columns=['primer_seq','size','start','len','tm','gc_percent','any_th',"3'_th","hairpin","seq",'adeptor_name','runname','sample_id','pos']
        dfnew['primer_name']=dfnew[['pos','adeptor_name']].apply(lambda x: '-'.join(x),axis=1)
        dfnew.index=['OLIGO_'+str(k)+'_LEFT','OLIGO_'+str(k)+'_RIGHT']
        dfnewnew=dfnew[['primer_name','primer_seq','primer_name','size','start','len','tm','gc_percent','any_th',"3'_th",'hairpin','seq','runname','sample_id','pos']]
        #print(dfnewnew)
        primer_result_pd=primer_result_pd.append(dfnewnew,ignore_index=True)
    print(primer_result_pd.to_string())
   
                # append pandas dataframe to a existing xlsx file
        #if k==0:
        #    dfnewnew.to_csv(output_txt, header=True, index=None, sep='\t', mode='a')
        #    append_df_to_excel(filepath,dfnewnew,index=False,sheet_name=sheetName)
        #if k!=0:
        #    dfnewnew.to_csv(output_txt, header=None, index=None, sep='\t', mode='a')
        #    append_df_to_excel(filepath,dfnewnew,header=None,index=False,sheet_name=sheetName)
    #output_txt.close()
    os.system('rm -rf '+paranew)
    return(primer_result_pd)

# ------------------------------------
# Main function
# ------------------------------------


def main():
    usage = "usage: %prog [options] -i "
    description = "automatically Sanger primer design version1.0. "

    optparser = OptionParser(version="%prog 0.1",description=description,usage=usage,add_help_option=False)
    optparser.add_option("-h","--help",action="help",help="Show this help message and exit.")
    #optparser.add_option("-s","--seq",dest="SEQ",type="string",action='store',
    #                     help="Input sequence for primer design ")
    optparser.add_option("-i","--input",dest="INPUT",type="string",action='store',
                         help="Input files for primer design provided by clinical team.")    
    optparser.add_option("-p","--para",dest="PARA",type="string",action='store',
                         help="Iuput PARAMETERS file.")    
    optparser.add_option("-n","--name",dest="NAME",type="string",action='store',
                         help="Name for this run.")
    optparser.add_option('-s',"--sequence",dest='SEQ',type='string',action='store',
                        help='Input fastq/fasta sequence data.')
    
    (options,args) = optparser.parse_args()   
    #SEQ = options.SEQ
    INPUT = options.INPUT
    NAME= options.NAME
    PARA = options.PARA
    SEQ = options.SEQ
    
    
    # Chesk input files
    if not PARA or not INPUT or not NAME or not SEQ :
        optparser.print_help()
        sys.exit(1)


    M13F='GTAAAACGACGGCCAG'
    M13R='CAGGAAACAGCTATGAC'
    adeptor=[M13F,M13R]
    # read input file

    #print(INPUT)
    data=pd.read_csv(INPUT,sep='\t')

    print(data.to_string())
    
    path=os.getcwd()
    result=[]
    for index, row in data.iterrows():
        print(index)
        chrom=row['CHROM']
        pos=row['POS']
        snpLen=len(row['REF'])

        # get genomic sequence by chrom:start,end; window_size=500
        window_size=500
        start=int(pos)-window_size
        end=int(pos)+window_size
        #print([chrom,start,end])
        seqs=fetchSequence(str(chrom),str(start-1),str(end),SEQ)
        #print(seqs)
        #sys.exit()
        seqs_input=seqs.split('\n')[1]
        #print(data.REF[0])
        seqs_output=seqs_input[0:500]+'['+seqs_input[500:(500+snpLen)]+']'+seqs_input[(500+snpLen):]
        #print(seqs_output)
        data.ix[index,'FLANK_SEQ']=seqs_output
        #print(seqs_input)
        data_sub=data.iloc[[index]]
        result=primerDesign2(window_size+1,snpLen,seqs_input,PARA,NAME,adeptor,data_sub)
        pathfile = path+'/'+NAME+'_primer_design_output.xlsx'
        pathfiletxt = path+'/'+NAME+'_primer_design_output.txt'
        if index==0:
            result.to_excel(pathfile,sheet_name=str(index),index=False)
            result.to_csv(pathfiletxt,index=None,sep='\t')
            continue
        if index!=0:
            book = load_workbook(pathfile)
            writer = pd.ExcelWriter(pathfile, engine = 'openpyxl')
            writer.book = book
            result.to_excel(writer,sheet_name=str(index),index=False)
            result.to_csv(pathfiletxt,index=None,sep='\t',mode='a',header=None)

            writer.save()
            writer.close()

    print(data.to_string())    
    data.to_csv(NAME+'_primer_design.flanks500.txt',index=None,sep='\t')
        
    
    
if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        sys.stderr.write("User interrupt me! ;-) See you!\n")
        sys.exit(0)       
